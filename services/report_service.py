from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from datetime import datetime
from models.report_model import ReportModel

BANNEDWORDS = ["LADDER", "SQUAD STRIKE", "TEAMS", "RANDOMS", "FULLYA SERIES", "FULL YA SERIES", "RAPIDILLO DE MASHING", "LABBIN' PALACE", "NOCTURNA"]


class ReportService:

    def __init__(self):
        self.model = ReportModel()

    def xlref(self, row, column):
        return get_column_letter(column) + str(row)

    def generate_excel(self, slug):
        idtorneo = self.model.get_event_id(slug)
        players = self.model.get_all_players(idtorneo)

        allsetplayer = {}
        playerinfo = {}

        for player in players:
            player_id = player['player']['id']
            allsetplayer[player_id] = self.model.get_sets_for_player(player_id)
            playerinfo[player_id] = player

        wb = Workbook()
        ws = wb.active

        ws.title = f"Clasheos"
        actualrow = 2

        ws.freeze_panes = ws['B2']
        colors = {
            'grey': PatternFill(start_color='525252', end_color='525252', fill_type="solid"),
            'yellow': PatternFill(start_color='cedb5a', end_color='cedb5a', fill_type="solid"),
            'red': PatternFill(start_color='f04d4d', end_color='f04d4d', fill_type="solid"),
            'green': PatternFill(start_color='59cf78', end_color='59cf78', fill_type="solid"),
        }
        today = datetime.today().date()
        ft = Font(color="000000", size=8)

        for player1id in allsetplayer:
            actualcolumn = 2
            ws.cell(row=actualrow, column=1, value=playerinfo[player1id]["gamerTag"]).font = ft

            for player2id in allsetplayer:
                ws.cell(row=1, column=actualcolumn, value=playerinfo[player2id]["gamerTag"]).font = ft

                if player1id != player2id:
                    countmax = 0
                    commentstr = ""

                    for set in allsetplayer[player1id]:
                        tournamentname = set["event"]["tournament"]["name"] + " " + set["event"]["name"]
                        skip = any(bannedword.lower() in tournamentname.lower() for bannedword in BANNEDWORDS)

                        if set["event"]["videogame"]["id"] != 1386:
                            skip = True

                        if not skip and len(set["slots"]) > 1 and countmax < 3:
                            slots_valid = all(
                                slot["entrant"] and len(slot["entrant"]["participants"]) == 1
                                for slot in set["slots"]
                            )
                            if slots_valid and any(
                                slot["entrant"]["participants"][0]["player"]["id"] == player2id for slot in set["slots"]
                            ):
                                dt_object = datetime.fromtimestamp(set["event"]["tournament"]["startAt"]).strftime("%Y/%m/%d")
                                commentstr += dt_object + ": " + tournamentname + "\n"

                                if countmax == 0:
                                    ws.cell(row=actualrow, column=actualcolumn, value=dt_object)
                                    days = (today - datetime.strptime(dt_object, '%Y/%m/%d').date()).days
                                    color = 'green'
                                    if days < 20: color = 'red'
                                    elif days < 60: color = 'yellow'
                                    ws[self.xlref(actualrow, actualcolumn)].fill = colors[color]
                                    ws[self.xlref(actualrow, actualcolumn)].font = ft
                                countmax += 1

                    if commentstr:
                        ws[self.xlref(actualrow, actualcolumn)].comment = Comment(commentstr, "<author>", 120, 300)
                else:
                    ws[self.xlref(actualrow, actualcolumn)].fill = colors['grey']

                actualcolumn += 1
            actualrow += 1

        try:
            nombre_torneo = slug.split('/')[1]
        except:
            nombre_torneo = slug

        filename = f"Clasheos - {nombre_torneo}.xlsx"
        wb.save(filename)
        return filename
